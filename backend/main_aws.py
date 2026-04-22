from fastapi import FastAPI, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import pandas as pd
import io
import os
import math
import json
import re
import uuid
from datetime import datetime

from dynamo import DTable, create_tables

try:
    create_tables()
except Exception as e:
    print(f"Table creation skipped: {e}")

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

ublox_tb = DTable("ublox")
sales_tb = DTable("sales")

# ==================== 유틸 ====================

def clean_value(v):
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if isinstance(v, pd.Timestamp):
        return v.strftime("%Y-%m-%d")
    return v

def to_float(v):
    if v is None:
        return None
    try:
        f = float(v)
        return None if math.isnan(f) or math.isinf(f) else f
    except (ValueError, TypeError):
        return None

def safe_str(v):
    if v is None:
        return None
    s = str(v)
    return None if s in ("nan", "None", "") else s


# ==================== Microchip 매칭 ====================

COLUMNS = [
    "고객코드", "믹스#", "Sales", "고객", "END", "PURCHASING", "PART#", "FAB2", "LT",
    "2023년", "2024년", "2025년", "2026년", "23~25추이", "25-26(w/BL)", "BLOG TTL",
    "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월",
    "믹스#(customer&part)",
]
MONTH_COLUMNS = ["3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월"]


def _s(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    s = str(v).strip()
    return s if s and s.lower() != "nan" else None


def _code_str(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def _parse_snapshot_date(sheet_name: str):
    m = re.search(r"(\d{6})$", sheet_name)
    if not m:
        return None
    s = m.group(1)
    try:
        return pd.Timestamp(2000 + int(s[:2]), int(s[2:4]), int(s[4:6]))
    except ValueError:
        return None


def _bucket_month(crd, snapshot_date):
    if not isinstance(crd, pd.Timestamp) or pd.isna(crd):
        return None
    if snapshot_date and crd < snapshot_date:
        if snapshot_date.year == 2026 and 3 <= snapshot_date.month <= 12:
            return snapshot_date.month
        return None
    if crd.year == 2026 and 3 <= crd.month <= 12:
        return crd.month
    return None


def build_matching_records(contents: bytes):
    """백록YYMMDD + 출고내역 + FAB2 → 출고기준(백록매칭) 포맷"""
    xls = pd.ExcelFile(io.BytesIO(contents), engine="openpyxl")

    shipment_sheet = None
    backlog_sheet = None
    fab2_sheet = None
    for name in xls.sheet_names:
        if name == "출고내역":
            shipment_sheet = name
        elif name.startswith("백록") and "피벗" not in name:
            backlog_sheet = name
        elif name == "FAB2":
            fab2_sheet = name

    if not shipment_sheet and not backlog_sheet:
        return None, None, None

    snapshot_date = _parse_snapshot_date(backlog_sheet) if backlog_sheet else None

    ship_agg = {}
    if shipment_sheet:
        df = pd.read_excel(xls, sheet_name=shipment_sheet, header=0)
        for _, row in df.iterrows():
            mix = _s(row.get("믹스#"))
            if not mix:
                continue
            qty = to_float(row.get("출고수량")) or 0
            date = row.get("출고일자")
            year = date.year if isinstance(date, pd.Timestamp) and not pd.isna(date) else None

            if mix not in ship_agg:
                ship_agg[mix] = {
                    "고객코드": _code_str(row.get("고객코드")),
                    "Sales": _s(row.get("담당자")),
                    "고객": _s(row.get("고객")),
                    "END": _s(row.get("END고객사명")),
                    "PURCHASING": _s(row.get("PURCHSING")),
                    "PART#": _s(row.get("품번")),
                    "yearly": {},
                }
            if year:
                ship_agg[mix]["yearly"][year] = ship_agg[mix]["yearly"].get(year, 0) + qty

    bl_agg = {}
    if backlog_sheet:
        df = pd.read_excel(xls, sheet_name=backlog_sheet, header=0)
        for _, row in df.iterrows():
            mix = _s(row.get("믹스"))
            if not mix:
                continue
            qty = to_float(row.get("Qty Due")) or 0
            month = _bucket_month(row.get("CRD"), snapshot_date)
            lt = to_float(row.get("Lead Time Weeks"))

            if mix not in bl_agg:
                bl_agg[mix] = {
                    "고객코드": _code_str(row.get("업체코드")),
                    "고객": _s(row.get("업체명")),
                    "END": _s(row.get("End Customer Name")),
                    "PURCHASING": _s(row.get("ODM/SubCon Name")),
                    "PART#": _s(row.get("Customer Part Number")),
                    "LT": lt,
                    "monthly": {},
                }
            else:
                if bl_agg[mix]["LT"] is None and lt is not None:
                    bl_agg[mix]["LT"] = lt
            if month:
                bl_agg[mix]["monthly"][month] = bl_agg[mix]["monthly"].get(month, 0) + qty

    fab2_map = {}
    if fab2_sheet:
        df = pd.read_excel(xls, sheet_name=fab2_sheet, header=0)
        for _, row in df.iterrows():
            pn = _s(row.get("PN"))
            if pn:
                fab2_map[pn] = _s(row.get("PCN Number")) or _s(row.get("Remark")) or "Y"

    records = []
    for mix in set(ship_agg) | set(bl_agg):
        ship = ship_agg.get(mix, {})
        bl = bl_agg.get(mix, {})
        yearly = ship.get("yearly", {})
        monthly = bl.get("monthly", {})

        part_no = ship.get("PART#") or bl.get("PART#")
        end = ship.get("END") or bl.get("END")

        blog_ttl = sum(monthly.values()) if monthly else 0
        y2025 = yearly.get(2025)
        y2026 = yearly.get(2026)
        if y2025 is None and y2026 is None and blog_ttl == 0:
            wbl = None
        else:
            wbl = (y2026 or 0) + blog_ttl - (y2025 or 0)

        rec = {
            "고객코드": ship.get("고객코드") or bl.get("고객코드"),
            "믹스#": mix,
            "Sales": ship.get("Sales"),
            "고객": ship.get("고객") or bl.get("고객"),
            "END": end,
            "PURCHASING": ship.get("PURCHASING") or bl.get("PURCHASING"),
            "PART#": part_no,
            "FAB2": fab2_map.get(part_no, "-") if part_no else "-",
            "LT": bl.get("LT"),
            "2023년": yearly.get(2023),
            "2024년": yearly.get(2024),
            "2025년": yearly.get(2025),
            "2026년": yearly.get(2026),
            "23~25추이": None,
            "25-26(w/BL)": wbl,
            "BLOG TTL": blog_ttl,
        }
        for i, col in enumerate(MONTH_COLUMNS, start=3):
            rec[col] = monthly.get(i, 0)
        rec["믹스#(customer&part)"] = (end or "") + (part_no or "")
        records.append(rec)

    records.sort(key=lambda r: ((r.get("PART#") or ""), (r.get("END") or "")))
    return "출고기준(백록매칭)", COLUMNS, records


@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...)):
    contents = await file.read()
    target_sheet, final_columns, records = build_matching_records(contents)

    if target_sheet is None:
        return {"error": "백록 또는 출고내역 시트를 찾을 수 없습니다."}
    if records is None:
        return {"error": "데이터를 파싱할 수 없습니다."}

    return {
        "sheet_name": target_sheet,
        "columns": final_columns,
        "data": records,
        "total_rows": len(records),
    }


# ==================== u-blox 백로그 ====================

UBLOX_DISPLAY_COLUMNS = [
    "Order Name", "Order No", "PO No Line Item", "Invoice Number", "Reference",
    "Account Name", "Account Number", "Reporting uB Office", "Order Status",
    "Type Number", "Frame Order", "Order Date", "Request Date", "Delivery Date",
    "전일 Delivery Date", "DELINQ", "Qty Ordered", "Qty Invoiced",
    "Price per unit", "Total Value", "End Customer", "End Customer No",
    "Project Owner", "Project Owner No",
]

UBLOX_COL_MAP = {
    0: "Order Name", 1: "Order No", 2: "PO No Line Item", 3: "Invoice Number",
    4: "Reference", 5: "Account Name", 6: "Account Number", 7: "Reporting uB Office",
    8: "Order Status", 9: "Type Number", 10: "Frame Order", 11: "Order Date",
    12: "Request Date", 13: "Delivery Date", 14: "전일 Delivery Date",
    15: "DELINQ", 16: "Qty Ordered", 17: "Qty Invoiced", 18: "Price per unit",
    19: "Total Value", 20: "End Customer", 21: "End Customer No",
    22: "Project Owner", 23: "Project Owner No",
}
FLOAT_COLS = {"DELINQ", "Qty Ordered", "Qty Invoiced", "Price per unit", "Total Value"}


@app.post("/api/ublox/upload")
async def upload_ublox(file: UploadFile = File(...)):
    contents = await file.read()
    xls = pd.ExcelFile(io.BytesIO(contents), engine="openpyxl")
    df = pd.read_excel(xls, sheet_name=0, header=None)

    header_row = 0
    for i in range(min(5, len(df))):
        if "Order Name" in [str(v).strip() for v in df.iloc[i].values if pd.notna(v)]:
            header_row = i
            break

    df = df.iloc[header_row + 1:].reset_index(drop=True).dropna(how="all").reset_index(drop=True)

    records = []
    for _, row in df.iterrows():
        rec = {}
        has_data = False
        for idx, col_name in UBLOX_COL_MAP.items():
            if idx >= len(row):
                continue
            v = clean_value(row.iloc[idx])
            if col_name in FLOAT_COLS:
                rec[col_name] = to_float(v)
            elif isinstance(v, pd.Timestamp):
                rec[col_name] = v.strftime("%Y-%m-%d")
            else:
                rec[col_name] = safe_str(v)
            if v is not None:
                has_data = True
        if has_data and rec.get("Order Name"):
            records.append(rec)

    # 버전 관리
    all_existing = ublox_tb.scan_all()
    version_nums = set()
    for item in all_existing:
        try:
            version_nums.add(int(item.get("upload_version", "0")))
        except ValueError:
            pass
    prev_version_num = max(version_nums) if version_nums else 0
    new_version_num = prev_version_num + 1
    prev_version = str(prev_version_num) if prev_version_num > 0 else None
    new_version = str(new_version_num)

    # 전일 비교
    prev_map = {}
    if prev_version and prev_version != "0":
        prev_items = ublox_tb.query("upload_version", prev_version)
        for item in prev_items:
            d = json.loads(item.get("data", "{}"))
            prev_map[d.get("Order Name", "")] = d

    # 저장 (order_name + 행번호로 unique SK)
    for i, r in enumerate(records):
        ublox_tb.put({
            "upload_version": new_version,
            "order_name": f"{r.get('Order Name', '')}_{i}",
            "data": json.dumps(r, ensure_ascii=False, default=str),
        })

    # 변경 비교
    for r in records:
        change = {"type": None, "changed_fields": []}
        oname = r.get("Order Name", "")
        if oname not in prev_map:
            if prev_map:
                change["type"] = "new"
        else:
            prev = prev_map[oname]
            changed = [c for c in ["Delivery Date", "Qty Ordered", "Price per unit", "Order Status"]
                       if str(r.get(c, "")) != str(prev.get(c, ""))]
            if changed:
                change["type"] = "modified"
                change["changed_fields"] = changed
            del prev_map[oname]
        r["_change"] = change

    deleted = []
    for oname, prev in prev_map.items():
        prev["_change"] = {"type": "deleted", "changed_fields": []}
        deleted.append(prev)

    return {
        "columns": UBLOX_DISPLAY_COLUMNS, "data": records, "deleted": deleted,
        "total_rows": len(records), "has_prev": prev_version is not None and prev_version != "0",
        "version": int(new_version), "upload_date": str(datetime.now().date()),
    }


@app.get("/api/ublox/data")
async def get_ublox_data():
    all_items = ublox_tb.scan_all()
    if not all_items:
        return {"data": [], "columns": UBLOX_DISPLAY_COLUMNS, "total_rows": 0}

    version_nums = sorted(set(int(item.get("upload_version", "0")) for item in all_items))
    latest = str(version_nums[-1])
    prev = str(version_nums[-2]) if len(version_nums) > 1 else None

    latest_items = [json.loads(i.get("data", "{}")) for i in all_items if i.get("upload_version") == latest]
    prev_map = {}
    if prev:
        for i in all_items:
            if i.get("upload_version") == prev:
                d = json.loads(i.get("data", "{}"))
                prev_map[d.get("Order Name", "")] = d

    deleted = []
    for r in latest_items:
        change = {"type": None, "changed_fields": []}
        oname = r.get("Order Name", "")
        if prev_map:
            if oname not in prev_map:
                change["type"] = "new"
            else:
                p = prev_map[oname]
                changed = [c for c in ["Delivery Date", "Qty Ordered", "Price per unit", "Order Status"]
                           if str(r.get(c, "")) != str(p.get(c, ""))]
                if changed:
                    change["type"] = "modified"
                    change["changed_fields"] = changed
                del prev_map[oname]
        r["_change"] = change

    for oname, p in prev_map.items():
        p["_change"] = {"type": "deleted", "changed_fields": []}
        deleted.append(p)

    return {
        "columns": UBLOX_DISPLAY_COLUMNS, "data": latest_items, "deleted": deleted,
        "total_rows": len(latest_items), "upload_date": str(datetime.now().date()),
        "version": int(latest), "has_prev": prev is not None,
    }


@app.get("/api/ublox/search/{type_number}")
async def search_ublox(type_number: str):
    all_items = ublox_tb.scan_all()
    versions = sorted(set(i.get("upload_version", "0") for i in all_items))
    if not versions:
        return {"data": [], "summary": None}
    latest = versions[-1]
    records = []
    for i in all_items:
        if i.get("upload_version") != latest:
            continue
        d = json.loads(i.get("data", "{}"))
        if type_number.lower() in str(d.get("Type Number", "")).lower():
            records.append(d)

    total_qty = sum(to_float(r.get("Qty Ordered")) or 0 for r in records)
    total_value = sum(to_float(r.get("Total Value")) or 0 for r in records)
    customers = list(set(r.get("End Customer") for r in records if r.get("End Customer")))

    return {
        "columns": UBLOX_DISPLAY_COLUMNS, "data": records, "total_rows": len(records),
        "summary": {"type_number": type_number, "total_qty": total_qty,
                     "total_value": total_value, "order_count": len(records), "customers": customers},
    }


@app.delete("/api/ublox/data")
async def reset_ublox():
    ublox_tb.delete_all()
    return {"deleted": "all"}


# ==================== 영업실적 ====================

FIXED_EXCHANGE_RATE = 1400
SALES_COLUMNS = [
    "구분", "MPN", "QTY", "DCPL($)", "매입금액($)", "SP($)", "매출금액($)",
    "매출환율", "SP(KRW)", "매출금액(KRW)", "GP($)", "GP%($)", "GP(KRW)", "GP%(KRW)",
    "담당자", "납품처", "거래처코드", "출고일자", "입고일", "Month",
]


def parse_remark(remark):
    if not remark or str(remark) == "nan":
        return None, None
    parts = str(remark).split("_")
    try:
        buy = float(parts[0])
    except (ValueError, IndexError):
        buy = None
    try:
        sell = float(parts[1]) if len(parts) > 1 else None
    except (ValueError, IndexError):
        sell = None
    return buy, sell


def parse_lot_date(lot_no):
    if not lot_no or str(lot_no) == "nan":
        return None
    parts = str(lot_no).split("_")
    if not parts:
        return None
    ds = parts[0].strip()
    if len(ds) == 6 and ds.isdigit():
        return f"20{ds[:2]}-{ds[2:4]}-{ds[4:6]}"
    return None


@app.post("/api/sales/upload")
async def upload_sales(file: UploadFile = File(...)):
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents), header=0)
    except Exception:
        df = pd.read_excel(io.BytesIO(contents), header=0, engine="xlrd")

    records = []
    for _, row in df.iterrows():
        qty = to_float(row.get("출고수량"))
        if not qty or qty == 0:
            continue
        dcpl, sp = parse_remark(row.get("비고(내역)"))
        inbound_date = parse_lot_date(row.get("LOT No."))
        foreign_price = to_float(row.get("외화단가"))
        if foreign_price and foreign_price > 0:
            sp = foreign_price
        exch_rate = to_float(row.get("환율"))
        if not exch_rate or exch_rate <= 1:
            exch_rate = FIXED_EXCHANGE_RATE

        buy_amt = round(dcpl * qty, 2) if dcpl else None
        sell_amt = round(sp * qty, 2) if sp else None
        sp_krw = round(sp * exch_rate, 2) if sp else None
        sell_amt_krw = round(sell_amt * exch_rate, 2) if sell_amt else None
        gp_usd = round(sell_amt - buy_amt, 2) if sell_amt and buy_amt else None
        gp_pct = round(gp_usd / sell_amt * 100, 2) if gp_usd and sell_amt else None
        gp_krw = round(gp_usd * exch_rate, 2) if gp_usd else None
        gp_pct_krw = round(gp_krw / sell_amt_krw * 100, 2) if gp_krw and sell_amt_krw else None

        ship_date = row.get("출고일자")
        ship_date = ship_date.strftime("%Y-%m-%d") if isinstance(ship_date, pd.Timestamp) else safe_str(ship_date)
        month_val = safe_str(row.get("출고년월"))
        if month_val:
            month_val = month_val.replace("/", "")
        vendor = safe_str(row.get("품목군")) or ""

        records.append({
            "구분": vendor, "MPN": safe_str(row.get("품번")) or "",
            "QTY": qty, "DCPL($)": dcpl, "매입금액($)": buy_amt,
            "SP($)": sp, "매출금액($)": sell_amt, "매출환율": exch_rate,
            "SP(KRW)": sp_krw, "매출금액(KRW)": sell_amt_krw,
            "GP($)": gp_usd, "GP%($)": gp_pct, "GP(KRW)": gp_krw, "GP%(KRW)": gp_pct_krw,
            "담당자": safe_str(row.get("담당자")) or "", "납품처": safe_str(row.get("고객")) or "",
            "거래처코드": safe_str(row.get("고객코드")) or "", "출고일자": ship_date,
            "입고일": inbound_date, "Month": month_val,
        })

    # DB 저장
    sales_tb.delete_all()
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    for i, r in enumerate(records):
        sales_tb.put({
            "batch_id": batch_id, "item_id": str(i),
            "data": json.dumps(r, ensure_ascii=False, default=str),
        })

    total_sales = sum(r.get("매출금액($)") or 0 for r in records)
    total_buy = sum(r.get("매입금액($)") or 0 for r in records)
    total_gp = sum(r.get("GP($)") or 0 for r in records)
    total_sales_krw = sum(r.get("매출금액(KRW)") or 0 for r in records)
    total_gp_krw = sum(r.get("GP(KRW)") or 0 for r in records)

    return {
        "columns": SALES_COLUMNS, "data": records, "total_rows": len(records),
        "summary": {
            "total_sales_usd": round(total_sales, 2), "total_buy_usd": round(total_buy, 2),
            "total_gp_usd": round(total_gp, 2),
            "total_gp_pct": round(total_gp / total_sales * 100, 2) if total_sales else 0,
            "total_sales_krw": round(total_sales_krw, 2), "total_gp_krw": round(total_gp_krw, 2),
            "total_gp_pct_krw": round(total_gp_krw / total_sales_krw * 100, 2) if total_sales_krw else 0,
        }, "saved_to_db": True,
    }


@app.get("/api/sales/data")
async def get_sales_data():
    items = sales_tb.scan_all()
    if not items:
        return {"data": [], "columns": SALES_COLUMNS, "total_rows": 0}
    records = [json.loads(i.get("data", "{}")) for i in items]
    total_sales = sum(r.get("매출금액($)") or 0 for r in records)
    total_buy = sum(r.get("매입금액($)") or 0 for r in records)
    total_gp = sum(r.get("GP($)") or 0 for r in records)
    total_sales_krw = sum(r.get("매출금액(KRW)") or 0 for r in records)
    total_gp_krw = sum(r.get("GP(KRW)") or 0 for r in records)
    return {
        "columns": SALES_COLUMNS, "data": records, "total_rows": len(records),
        "summary": {
            "total_sales_usd": round(total_sales, 2), "total_buy_usd": round(total_buy, 2),
            "total_gp_usd": round(total_gp, 2),
            "total_gp_pct": round(total_gp / total_sales * 100, 2) if total_sales else 0,
            "total_sales_krw": round(total_sales_krw, 2), "total_gp_krw": round(total_gp_krw, 2),
            "total_gp_pct_krw": round(total_gp_krw / total_sales_krw * 100, 2) if total_sales_krw else 0,
        },
    }


# ==================== 마이크론 재고 ====================

MICRON_COLUMNS = [
    "Status", "Type", "PO", "DID", "MPN", "CPN (MOBIS ID 포함)", "BOX_TYPE",
    "QTY", "DNNo.", "Ship Date", "MicronInvoice#", "수입면장번호", "BL번호",
    "수입신고일", "FSE", "End customer", "Date Code",
    "Booking Customer & FSE", "Qty_booking", "비고",
]

micron_data = []


@app.post("/api/micron/upload")
async def upload_micron(file: UploadFile = File(...)):
    global micron_data
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents), sheet_name="Detail", header=0)
    except Exception:
        return {"error": "Detail 시트를 찾을 수 없습니다."}

    records = []
    for idx, row in df.iterrows():
        rec = {}
        for col in df.columns:
            v = row[col]
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                rec[col] = None
            elif isinstance(v, pd.Timestamp):
                rec[col] = v.strftime("%Y-%m-%d")
            elif v is not None:
                rec[col] = str(v) if not isinstance(v, (int, float)) else v
            else:
                rec[col] = None
        rec["_id"] = str(idx)
        records.append(rec)

    micron_data = records
    return {"columns": MICRON_COLUMNS, "data": records, "total_rows": len(records)}


@app.get("/api/micron/data")
async def get_micron_data(status: str = "", did: str = "", mpn: str = "", notes_only: str = ""):
    filtered = micron_data
    if status:
        filtered = [r for r in filtered if status in str(r.get("Status", ""))]
    if did:
        filtered = [r for r in filtered if did.lower() in str(r.get("DID", "")).lower()]
    if mpn:
        filtered = [r for r in filtered if mpn.lower() in str(r.get("MPN", "")).lower()]
    if notes_only == "true":
        filtered = [r for r in filtered if r.get("비고") and str(r.get("비고")) not in ("None", "", "nan")]
    return {"columns": MICRON_COLUMNS, "data": filtered, "total_rows": len(filtered)}


@app.get("/api/micron/summary/{did}")
async def micron_summary(did: str):
    items = [r for r in micron_data if str(r.get("DID", "")).upper() == did.upper()]
    if not items:
        return {"error": "해당 DID 없음"}
    by_status = {}
    for r in items:
        s = str(r.get("Status", "기타"))
        if s not in by_status:
            by_status[s] = {"count": 0, "qty": 0}
        by_status[s]["count"] += 1
        by_status[s]["qty"] += to_float(r.get("QTY")) or 0
    total_qty = sum(v["qty"] for v in by_status.values())
    mpns = list(set(str(r.get("MPN", "")) for r in items))
    return {"did": did, "mpns": mpns, "total_qty": total_qty, "by_status": by_status, "items": items}


@app.post("/api/micron/update")
async def update_micron(request: Request):
    data = await request.json()
    item_id = data.get("_id")
    if item_id is None:
        return {"error": "ID 없음"}
    for r in micron_data:
        if str(r.get("_id")) == str(item_id):
            r["Booking Customer & FSE"] = data.get("Booking Customer & FSE", r.get("Booking Customer & FSE"))
            r["Qty_booking"] = data.get("Qty_booking", r.get("Qty_booking"))
            r["비고"] = data.get("비고", r.get("비고"))
            return {"updated": item_id}
    return {"error": "not found"}


# ==================== 엑셀 내보내기 ====================

@app.post("/api/export")
async def export_excel(request: Request):
    from openpyxl.styles import PatternFill, Font, Alignment
    data = await request.json()
    rows = data.get("data", [])
    columns = data.get("columns", [])
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="마이크로칩(매칭)", index=False)
        ws = writer.sheets["마이크로칩(매칭)"]
        sky_blue = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = sky_blue
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.auto_filter.ref = ws.dimensions
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=export.xlsx"})


# ==================== 환율 + 거래명세서 ====================

import requests as http_requests

@app.get("/api/exchange-rate")
async def get_exchange_rate():
    try:
        r = http_requests.get("https://open.er-api.com/v6/latest/USD", timeout=5)
        rate = r.json()["rates"]["KRW"]
        return {"rate": round(rate, 2), "source": "exchangerate-api"}
    except Exception:
        return {"rate": 1400, "source": "fallback"}


@app.post("/api/invoice/generate")
async def generate_invoice(request: Request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

    data = await request.json()
    items = data.get("items", [])
    customer = data.get("customer", "")
    date_str = data.get("date", "")
    rate = float(data.get("rate", 1400))

    wb = Workbook()
    ws = wb.active
    ws.title = "거래명세서"

    widths = {"A": 5, "B": 20, "C": 8, "D": 14, "E": 16, "F": 10, "G": 16, "H": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    thin = Side(style="thin")
    thick = Side(style="medium")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left", vertical="center")

    # 제목
    ws.merge_cells("A1:H1")
    ws["A1"] = "거래명세표"
    ws["A1"].font = Font(bold=True, size=36)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 55
    ws.row_dimensions[2].height = 8

    # 공급자
    ws.merge_cells("A3:C3")
    ws["A3"] = "공 급 자"
    ws["A3"].font = Font(bold=True, size=10)
    ws["A3"].alignment = center
    ws["A3"].border = Border(left=thick, top=thick, bottom=thin)
    ws["B3"].border = Border(top=thick, bottom=thin)
    ws["C3"].border = Border(right=thick, top=thick, bottom=thin)

    info = [(4, "등록번호 : 229-81-00105"), (5, "상      호 : ㈜유니트론텍"),
            (6, "대표이사 : 남궁 선"), (7, "주 : 서울 강남구 영동대로 638(삼성동, 삼보빌딩 9층)"),
            (8, "업      태 : 도.소매"), (9, "종      목 :전자부품 외")]
    for row_num, val in info:
        ws.merge_cells(f"A{row_num}:C{row_num}")
        ws[f"A{row_num}"] = val
        ws[f"A{row_num}"].font = Font(size=8)
        ws[f"A{row_num}"].alignment = left_a
        if row_num == 9:
            ws[f"A{row_num}"].border = Border(left=thick, bottom=thick)
            ws[f"B{row_num}"].border = Border(bottom=thick)
            ws[f"C{row_num}"].border = Border(right=thick, bottom=thick)
        else:
            ws[f"A{row_num}"].border = Border(left=thick)
            ws[f"C{row_num}"].border = Border(right=thick)

    # 공급받는자
    ws.merge_cells("F3:H3")
    ws["F3"] = "공급받는자"
    ws["F3"].font = Font(bold=True, size=10)
    ws["F3"].alignment = center
    ws["F3"].border = Border(left=thick, top=thick, bottom=thin)
    ws["G3"].border = Border(top=thick, bottom=thin)
    ws["H3"].border = Border(right=thick, top=thick, bottom=thin)

    ws.merge_cells("F4:H9")
    ws["F4"] = customer
    ws["F4"].font = Font(bold=True, size=16)
    ws["F4"].alignment = center
    for rn in range(4, 10):
        for cl in ["F", "G", "H"]:
            l = thick if cl == "F" else Side()
            r = thick if cl == "H" else Side()
            b = thick if rn == 9 else Side()
            ws[f"{cl}{rn}"].border = Border(left=l, right=r, bottom=b)

    ws["B12"] = date_str
    ws["B12"].font = Font(bold=True, size=10)

    # 헤더
    headers = ["No.", "Part #", "QTY", "U/PRICE ($)", "Amount ($)", "RATE", "U/PRICE (￦)", "AMOUNT (￦)"]
    header_fill = PatternFill(start_color="CCFF33", end_color="CCFF33", fill_type="solid")
    for j, h in enumerate(headers):
        cell = ws.cell(row=13, column=j+1)
        cell.value = h
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    total_usd = total_krw = 0
    for i, item in enumerate(items):
        row = 14 + i
        qty = float(item.get("qty", 0))
        price_usd = float(item.get("price", 0))
        amount_usd = round(qty * price_usd, 2)
        price_krw = round(price_usd * rate, 2)
        amount_krw = round(amount_usd * rate, 0)
        total_usd += amount_usd
        total_krw += amount_krw

        vals = [i+1, item.get("part",""), int(qty), price_usd, amount_usd, rate, price_krw, int(amount_krw)]
        fmts = [None, None, None, '$#,##0.00', '$#,##0.00', '#,##0.00', '₩#,##0.00', '₩#,##0']
        for j, (v, fmt) in enumerate(zip(vals, fmts)):
            c = ws.cell(row=row, column=j+1, value=v)
            c.alignment = center
            c.border = border
            if fmt:
                c.number_format = fmt

    for i in range(len(items), 10):
        row = 14 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center
        for j in range(1, 9):
            ws.cell(row=row, column=j).border = border

    sr = 24
    tax_usd = round(total_usd * 0.1, 2)
    tax_krw = round(total_krw * 0.1, 0)
    for label, uv, kv, off in [("소  계", total_usd, total_krw, 0), ("부가세", tax_usd, tax_krw, 1),
                                ("합  계", total_usd+tax_usd, total_krw+tax_krw, 2)]:
        r = sr + off
        fnt = Font(bold=True, size=9)
        ws.cell(row=r, column=4, value=label).font = fnt
        ws.cell(row=r, column=4).alignment = center
        ws.cell(row=r, column=5, value=uv).number_format = '$#,##0.00'
        ws.cell(row=r, column=5).alignment = center
        ws.cell(row=r, column=5).font = fnt
        ws.cell(row=r, column=7, value=label).font = fnt
        ws.cell(row=r, column=7).alignment = center
        ws.cell(row=r, column=8, value=int(kv)).number_format = '₩#,##0'
        ws.cell(row=r, column=8).alignment = center
        ws.cell(row=r, column=8).font = fnt
        for j in [4, 5, 7, 8]:
            ws.cell(row=r, column=j).border = border

    # 합계 아래 줄
    for j in range(1, 9):
        c = ws.cell(row=sr+2, column=j)
        c.border = Border(left=c.border.left, right=c.border.right, top=c.border.top, bottom=Side(style="medium"))

    ws.merge_cells("A27:H27")
    ws["A27"] = "  비  고 : 금일 최초고시 매매기준율 적용"
    ws["A27"].font = Font(size=9)

    ws.merge_cells("E30:H30")
    ws["E30"] = "인수자 :                                              (인)"
    for j in range(1, 9):
        ws.cell(row=30, column=j).border = Border(bottom=thin)

    ws["B31"] = "계좌정보"
    ws["B31"].font = Font(bold=True, size=9)
    ws["B32"] = "원화> 기업은행 528-002245-01011"
    ws["B32"].font = Font(size=9)
    ws["B33"] = "외화> 기업은행 528-002245-56-00013"
    ws["B33"].font = Font(size=9)

    stamp_path = os.path.join(os.path.dirname(__file__), "stamp.png")
    if os.path.exists(stamp_path):
        img = XlImage(stamp_path)
        img.width = 75
        img.height = 75
        m1 = AnchorMarker(col=2, colOff=300000, row=1, rowOff=50000)
        m2 = AnchorMarker(col=3, colOff=200000, row=4, rowOff=100000)
        img.anchor = TwoCellAnchor(_from=m1, to=m2)
        ws.add_image(img)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=invoice_{date_str}.xlsx"})


@app.post("/api/reset-tables")
async def reset_tables():
    ublox_tb.delete_all()
    sales_tb.delete_all()
    return {"status": "ok"}


# 프론트엔드 정적 파일 서빙
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=os.path.join(STATIC_DIR, "static")), name="static-assets")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        file_path = os.path.join(STATIC_DIR, full_path)
        if full_path and os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(STATIC_DIR, "index.html"))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
