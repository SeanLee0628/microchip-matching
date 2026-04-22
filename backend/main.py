from fastapi import FastAPI, UploadFile, File, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
import pandas as pd
import io
import os
import math
import re
import uuid
from datetime import datetime

from database import engine, get_db, Base
from models_ublox import UbloxBacklog, UBLOX_COLUMN_MAP, UBLOX_DISPLAY_COLUMNS
from models_sales import SalesPerformance as SalesModel, SALES_FIELD_MAP, SALES_REVERSE_MAP

# 테이블 생성
Base.metadata.create_all(bind=engine)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

COLUMNS = [
    "고객코드", "믹스#", "Sales", "고객", "END", "PURCHASING", "PART#", "FAB2", "LT",
    "2023년", "2024년", "2025년", "2026년", "23~25추이", "25-26(w/BL)", "BLOG TTL",
    "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월",
    "믹스#(customer&part)",
]

MONTH_COLUMNS = ["3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월"]


def clean_value(v):
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if isinstance(v, pd.Timestamp):
        return v.strftime("%Y-%m-%d")
    return v


def to_float(v):
    """DB 저장용 float 변환"""
    if v is None:
        return None
    try:
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return None
        return f
    except (ValueError, TypeError):
        return None


def _s(v):
    """str 변환, NaN/None은 None"""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    s = str(v).strip()
    return s if s and s.lower() != "nan" else None


def _code_str(v):
    """고객코드: float(131112.0) → '131112'"""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def _parse_snapshot_date(sheet_name: str):
    """'백록260324' → Timestamp(2026, 3, 24)"""
    m = re.search(r"(\d{6})$", sheet_name)
    if not m:
        return None
    s = m.group(1)
    try:
        return pd.Timestamp(2000 + int(s[:2]), int(s[2:4]), int(s[4:6]))
    except ValueError:
        return None


def _bucket_month(crd, snapshot_date):
    """CRD를 2026년 월 버킷에 배치. snapshot 이전이면 snapshot 월로 당김."""
    if not isinstance(crd, pd.Timestamp) or pd.isna(crd):
        return None
    if snapshot_date and crd < snapshot_date:
        if snapshot_date.year == 2026 and 3 <= snapshot_date.month <= 12:
            return snapshot_date.month
        return None
    if crd.year == 2026 and 3 <= crd.month <= 12:
        return crd.month
    return None


def parse_excel(contents: bytes):
    """백록260324 + 출고내역 + FAB2 → 출고기준(백록매칭) 포맷 레코드 조립"""
    xls = pd.ExcelFile(io.BytesIO(contents), engine="openpyxl")

    shipment_sheet = None
    backlog_sheet = None
    fab2_sheet = None
    for name in xls.sheet_names:
        if name == "출고내역":
            shipment_sheet = name
        elif name.startswith("백록") and "피벗" not in name:
            backlog_sheet = name
        elif name == "FAB2":
            fab2_sheet = name

    if not shipment_sheet and not backlog_sheet:
        return None, None, None

    snapshot_date = _parse_snapshot_date(backlog_sheet) if backlog_sheet else None

    # 출고내역 집계: 믹스# → 고객/PART 정보 + 연도별 출고수량 합계
    ship_agg = {}
    if shipment_sheet:
        df = pd.read_excel(xls, sheet_name=shipment_sheet, header=0)
        for _, row in df.iterrows():
            mix = _s(row.get("믹스#"))
            if not mix:
                continue
            qty = to_float(row.get("출고수량")) or 0
            date = row.get("출고일자")
            year = date.year if isinstance(date, pd.Timestamp) and not pd.isna(date) else None

            if mix not in ship_agg:
                ship_agg[mix] = {
                    "고객코드": _code_str(row.get("고객코드")),
                    "Sales": _s(row.get("담당자")),
                    "고객": _s(row.get("고객")),
                    "END": _s(row.get("END고객사명")),
                    "PURCHASING": _s(row.get("PURCHSING")),
                    "PART#": _s(row.get("품번")),
                    "yearly": {},
                }
            if year:
                ship_agg[mix]["yearly"][year] = ship_agg[mix]["yearly"].get(year, 0) + qty

    # 백록 집계: 믹스 → LT + 월별 Qty Due 합계
    bl_agg = {}
    if backlog_sheet:
        df = pd.read_excel(xls, sheet_name=backlog_sheet, header=0)
        for _, row in df.iterrows():
            mix = _s(row.get("믹스"))
            if not mix:
                continue
            qty = to_float(row.get("Qty Due")) or 0
            month = _bucket_month(row.get("CRD"), snapshot_date)
            lt = to_float(row.get("Lead Time Weeks"))

            if mix not in bl_agg:
                bl_agg[mix] = {
                    "고객코드": _code_str(row.get("업체코드")),
                    "고객": _s(row.get("업체명")),
                    "END": _s(row.get("End Customer Name")),
                    "PURCHASING": _s(row.get("ODM/SubCon Name")),
                    "PART#": _s(row.get("Customer Part Number")),
                    "LT": lt,
                    "monthly": {},
                }
            else:
                # LT가 비어있던 행이 있으면 보충
                if bl_agg[mix]["LT"] is None and lt is not None:
                    bl_agg[mix]["LT"] = lt
            if month:
                bl_agg[mix]["monthly"][month] = bl_agg[mix]["monthly"].get(month, 0) + qty

    # FAB2 매핑: PART# → PCN 마킹
    fab2_map = {}
    if fab2_sheet:
        df = pd.read_excel(xls, sheet_name=fab2_sheet, header=0)
        for _, row in df.iterrows():
            pn = _s(row.get("PN"))
            if pn:
                fab2_map[pn] = _s(row.get("PCN Number")) or _s(row.get("Remark")) or "Y"

    # 병합
    records = []
    for mix in set(ship_agg) | set(bl_agg):
        ship = ship_agg.get(mix, {})
        bl = bl_agg.get(mix, {})
        yearly = ship.get("yearly", {})
        monthly = bl.get("monthly", {})

        part_no = ship.get("PART#") or bl.get("PART#")
        end = ship.get("END") or bl.get("END")

        blog_ttl = sum(monthly.values()) if monthly else 0
        y2025 = yearly.get(2025)
        y2026 = yearly.get(2026)
        if y2025 is None and y2026 is None and blog_ttl == 0:
            wbl = None
        else:
            wbl = (y2026 or 0) + blog_ttl - (y2025 or 0)

        rec = {
            "고객코드": ship.get("고객코드") or bl.get("고객코드"),
            "믹스#": mix,
            "Sales": ship.get("Sales"),
            "고객": ship.get("고객") or bl.get("고객"),
            "END": end,
            "PURCHASING": ship.get("PURCHASING") or bl.get("PURCHASING"),
            "PART#": part_no,
            "FAB2": fab2_map.get(part_no, "-") if part_no else "-",
            "LT": bl.get("LT"),
            "2023년": yearly.get(2023),
            "2024년": yearly.get(2024),
            "2025년": yearly.get(2025),
            "2026년": yearly.get(2026),
            "23~25추이": None,
            "25-26(w/BL)": wbl,
            "BLOG TTL": blog_ttl,
        }
        for i, col in enumerate(MONTH_COLUMNS, start=3):
            rec[col] = monthly.get(i, 0)
        rec["믹스#(customer&part)"] = (end or "") + (part_no or "")
        records.append(rec)

    records.sort(key=lambda r: ((r.get("PART#") or ""), (r.get("END") or "")))

    return "출고기준(백록매칭)", COLUMNS, records


@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...)):
    contents = await file.read()
    target_sheet, final_columns, records = parse_excel(contents)

    if target_sheet is None:
        return {"error": "백록 또는 출고내역 시트를 찾을 수 없습니다."}
    if records is None:
        return {"error": "데이터를 파싱할 수 없습니다."}

    return {
        "sheet_name": target_sheet,
        "columns": final_columns,
        "data": records,
        "total_rows": len(records),
    }


@app.post("/api/export")
async def export_excel(data: dict):
    from openpyxl.styles import PatternFill, Font, Alignment

    rows = data.get("data", [])
    columns = data.get("columns", COLUMNS)

    df = pd.DataFrame(rows, columns=[c for c in columns if c in (rows[0].keys() if rows else columns)])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="마이크로칩(매칭)", index=False)
        ws = writer.sheets["마이크로칩(매칭)"]

        # 헤더 스타일: 하늘색 배경 + 볼드 + 가운데 정렬
        sky_blue = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center")

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = sky_blue
            cell.font = bold_font
            cell.alignment = center_align

        # 필터 설정
        ws.auto_filter.ref = ws.dimensions

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=microchip_matching_export.xlsx"},
    )


# ==================== u-blox 백로그 ====================

def parse_ublox_excel(contents: bytes):
    """u-blox 백로그 엑셀 파싱"""
    xls = pd.ExcelFile(io.BytesIO(contents), engine="openpyxl")
    df = pd.read_excel(xls, sheet_name=0, header=None)

    # 헤더 행 찾기
    header_row = 0
    for i in range(min(5, len(df))):
        row_values = [str(v).strip() for v in df.iloc[i].values if pd.notna(v)]
        if "Order Name" in row_values or "Type Number" in row_values:
            header_row = i
            break

    # 14번 컬럼의 헤더 (전일 날짜)
    prev_date_header = str(df.iloc[header_row, 14]) if df.shape[1] > 14 else None

    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df = df.dropna(how="all").reset_index(drop=True)

    records = []
    for _, row in df.iterrows():
        record = {}
        has_data = False
        for col_idx, (db_field, display_name, dtype) in UBLOX_COLUMN_MAP.items():
            if col_idx >= len(row):
                record[display_name] = None
                continue
            v = row.iloc[col_idx]
            v = clean_value(v)
            if dtype == "float":
                record[display_name] = to_float(v)
            elif dtype == "date" and isinstance(v, str):
                record[display_name] = v
            elif v is not None and isinstance(v, pd.Timestamp):
                record[display_name] = v.strftime("%Y-%m-%d")
            else:
                record[display_name] = str(v) if v is not None else None
            if v is not None:
                has_data = True
        if has_data and record.get("Order Name"):
            records.append(record)

    return records, prev_date_header


def ublox_record_to_db(record: dict, upload_date, version: int) -> UbloxBacklog:
    kwargs = {"upload_date": upload_date, "upload_version": version}
    for col_idx, (db_field, display_name, dtype) in UBLOX_COLUMN_MAP.items():
        val = record.get(display_name)
        if dtype == "float":
            kwargs[db_field] = to_float(val)
        else:
            kwargs[db_field] = str(val) if val is not None else None
    return UbloxBacklog(**kwargs)


def ublox_db_to_record(row: UbloxBacklog) -> dict:
    record = {}
    for col_idx, (db_field, display_name, dtype) in UBLOX_COLUMN_MAP.items():
        record[display_name] = getattr(row, db_field)
    return record


@app.post("/api/ublox/upload")
async def upload_ublox(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    records, prev_date_header = parse_ublox_excel(contents)

    if not records:
        return {"error": "데이터를 찾을 수 없습니다."}

    today = datetime.now().date()

    # 최신 버전 번호 조회
    latest_version = db.query(UbloxBacklog.upload_version).order_by(
        UbloxBacklog.upload_version.desc()
    ).first()
    prev_version = latest_version[0] if latest_version else None
    new_version = (prev_version or 0) + 1

    # 이전 데이터 조회 (직전 업로드와 비교)
    prev_records_map = {}
    if prev_version:
        prev_rows = db.query(UbloxBacklog).filter(
            UbloxBacklog.upload_version == prev_version
        ).all()
        for r in prev_rows:
            key = r.order_name
            prev_records_map[key] = ublox_db_to_record(r)

    # 새 버전으로 저장
    for r in records:
        db.add(ublox_record_to_db(r, today, new_version))
    db.commit()

    # 변경 비교
    changes = []
    for r in records:
        order_name = r.get("Order Name")
        change_info = {"type": None, "changed_fields": []}

        if order_name not in prev_records_map:
            if prev_records_map:  # 전일 데이터가 있을 때만 신규 표시
                change_info["type"] = "new"
        else:
            prev = prev_records_map[order_name]
            changed = []
            for col in ["Delivery Date", "Qty Ordered", "Price per unit", "Order Status"]:
                if str(r.get(col, "")) != str(prev.get(col, "")):
                    changed.append(col)
            if changed:
                change_info["type"] = "modified"
                change_info["changed_fields"] = changed
            del prev_records_map[order_name]

        r["_change"] = change_info

    # 삭제된 주문
    deleted = []
    for order_name, prev in prev_records_map.items():
        prev["_change"] = {"type": "deleted", "changed_fields": []}
        deleted.append(prev)

    return {
        "columns": UBLOX_DISPLAY_COLUMNS,
        "data": records,
        "deleted": deleted,
        "total_rows": len(records),
        "has_prev": prev_version is not None,
        "prev_version": prev_version,
        "version": new_version,
        "upload_date": str(today),
    }


@app.get("/api/ublox/data")
async def get_ublox_data(db: Session = Depends(get_db)):
    """최신 u-blox 데이터 조회 (전일 비교 포함)"""
    versions = db.query(UbloxBacklog.upload_version).distinct().order_by(
        UbloxBacklog.upload_version.desc()
    ).limit(2).all()

    if not versions:
        return {"data": [], "columns": UBLOX_DISPLAY_COLUMNS, "total_rows": 0}

    latest_version = versions[0][0]
    prev_version = versions[1][0] if len(versions) > 1 else None

    rows = db.query(UbloxBacklog).filter(
        UbloxBacklog.upload_version == latest_version
    ).order_by(UbloxBacklog.id).all()

    records = [ublox_db_to_record(r) for r in rows]

    # 전일 비교
    prev_records_map = {}
    deleted = []
    if prev_version:
        prev_rows = db.query(UbloxBacklog).filter(
            UbloxBacklog.upload_version == prev_version
        ).all()
        for r in prev_rows:
            prev_records_map[r.order_name] = ublox_db_to_record(r)

        for r in records:
            order_name = r.get("Order Name")
            change_info = {"type": None, "changed_fields": []}
            if order_name not in prev_records_map:
                change_info["type"] = "new"
            else:
                prev = prev_records_map[order_name]
                changed = []
                for col in ["Delivery Date", "Qty Ordered", "Price per unit", "Order Status"]:
                    if str(r.get(col, "")) != str(prev.get(col, "")):
                        changed.append(col)
                if changed:
                    change_info["type"] = "modified"
                    change_info["changed_fields"] = changed
                del prev_records_map[order_name]
            r["_change"] = change_info

        for order_name, prev in prev_records_map.items():
            prev["_change"] = {"type": "deleted", "changed_fields": []}
            deleted.append(prev)
    else:
        for r in records:
            r["_change"] = {"type": None, "changed_fields": []}

    return {
        "columns": UBLOX_DISPLAY_COLUMNS,
        "data": records,
        "deleted": deleted,
        "total_rows": len(records),
        "upload_date": str(rows[0].upload_date) if rows else None,
        "version": latest_version,
        "has_prev": prev_version is not None,
    }


@app.get("/api/ublox/search/{type_number}")
async def search_ublox(type_number: str, db: Session = Depends(get_db)):
    """품명으로 백로그 조회"""
    latest_version = db.query(UbloxBacklog.upload_version).order_by(
        UbloxBacklog.upload_version.desc()
    ).first()

    if not latest_version:
        return {"data": [], "summary": None}

    rows = db.query(UbloxBacklog).filter(
        UbloxBacklog.upload_version == latest_version[0],
        UbloxBacklog.type_number.ilike(f"%{type_number}%")
    ).order_by(UbloxBacklog.request_date).all()

    records = [ublox_db_to_record(r) for r in rows]

    # 요약
    total_qty = sum(to_float(r.get("Qty Ordered")) or 0 for r in records)
    total_value = sum(to_float(r.get("Total Value")) or 0 for r in records)
    customers = list(set(r.get("End Customer") for r in records if r.get("End Customer")))

    return {
        "columns": UBLOX_DISPLAY_COLUMNS,
        "data": records,
        "total_rows": len(records),
        "summary": {
            "type_number": type_number,
            "total_qty": total_qty,
            "total_value": total_value,
            "order_count": len(records),
            "customers": customers,
        },
    }


@app.delete("/api/ublox/data")
async def reset_ublox(db: Session = Depends(get_db)):
    count = db.query(UbloxBacklog).count()
    db.query(UbloxBacklog).delete()
    db.commit()
    return {"deleted": count}


# ==================== 영업실적 ====================

FIXED_EXCHANGE_RATE = 1400

SALES_COLUMNS = [
    "구분", "MPN", "QTY", "DCPL($)", "매입금액($)",
    "SP($)", "매출금액($)", "매출환율", "SP(KRW)", "매출금액(KRW)",
    "GP($)", "GP%($)", "GP(KRW)", "GP%(KRW)",
    "담당자", "납품처", "거래처코드", "출고일자", "입고일", "Month",
]


def parse_remark(remark):
    """비고(내역)에서 매입단가, 매출단가 추출: '2.2_2.8' → (2.2, 2.8)"""
    if not remark or str(remark) == "nan":
        return None, None
    parts = str(remark).split("_")
    try:
        buy = float(parts[0])
    except (ValueError, IndexError):
        buy = None
    try:
        sell = float(parts[1]) if len(parts) > 1 else None
    except (ValueError, IndexError):
        sell = None
    return buy, sell


def parse_lot_date(lot_no):
    """LOT No.에서 입고일 추출: '260129_8K_$2.2_Korni' → '2026-01-29'"""
    if not lot_no or str(lot_no) == "nan":
        return None
    parts = str(lot_no).split("_")
    if not parts:
        return None
    date_str = parts[0].strip()
    if len(date_str) == 6 and date_str.isdigit():
        yy, mm, dd = date_str[:2], date_str[2:4], date_str[4:6]
        return f"20{yy}-{mm}-{dd}"
    return None


def sales_record_to_db(record: dict, batch_id: str) -> SalesModel:
    float_fields = {"QTY", "DCPL($)", "매입금액($)", "SP($)", "매출금액($)", "매출환율",
                     "SP(KRW)", "매출금액(KRW)", "GP($)", "GP%($)", "GP(KRW)", "GP%(KRW)"}
    kwargs = {"upload_batch": batch_id}
    for excel_col, db_field in SALES_FIELD_MAP.items():
        val = record.get(excel_col)
        if excel_col in float_fields:
            kwargs[db_field] = to_float(val)
        else:
            kwargs[db_field] = str(val) if val is not None else None
    return SalesModel(**kwargs)


def sales_db_to_record(row: SalesModel) -> dict:
    record = {}
    for excel_col, db_field in SALES_FIELD_MAP.items():
        record[excel_col] = getattr(row, db_field)
    return record


def compute_sales_summary(records):
    total_sales_usd = sum(r.get("매출금액($)") or 0 for r in records)
    total_buy_usd = sum(r.get("매입금액($)") or 0 for r in records)
    total_gp_usd = sum(r.get("GP($)") or 0 for r in records)
    total_sales_krw = sum(r.get("매출금액(KRW)") or 0 for r in records)
    total_gp_krw = sum(r.get("GP(KRW)") or 0 for r in records)
    return {
        "total_sales_usd": round(total_sales_usd, 2),
        "total_buy_usd": round(total_buy_usd, 2),
        "total_gp_usd": round(total_gp_usd, 2),
        "total_gp_pct": round(total_gp_usd / total_sales_usd * 100, 2) if total_sales_usd else 0,
        "total_sales_krw": round(total_sales_krw, 2),
        "total_gp_krw": round(total_gp_krw, 2),
        "total_gp_pct_krw": round(total_gp_krw / total_sales_krw * 100, 2) if total_sales_krw else 0,
    }


@app.post("/api/sales/upload")
async def upload_sales(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents), header=0)
    except Exception:
        df = pd.read_excel(io.BytesIO(contents), header=0, engine="xlrd")

    records = []
    for _, row in df.iterrows():
        remark = row.get("비고(내역)")
        lot_no = row.get("LOT No.")
        qty = to_float(row.get("출고수량"))
        if not qty or qty == 0:
            continue

        dcpl, sp = parse_remark(remark)
        inbound_date = parse_lot_date(lot_no)

        # 외화단가가 있으면 SP로 사용 (USD 거래)
        foreign_price = to_float(row.get("외화단가"))
        if foreign_price and foreign_price > 0:
            sp = foreign_price

        # 매출환율: 더존 환율이 있으면 사용, 없으면 고정
        exch_rate = to_float(row.get("환율"))
        if not exch_rate or exch_rate <= 1:
            exch_rate = FIXED_EXCHANGE_RATE

        # 계산
        buy_amt = round(dcpl * qty, 2) if dcpl else None
        sell_amt = round(sp * qty, 2) if sp else None
        sp_krw = round(sp * exch_rate, 2) if sp else None
        sell_amt_krw = round(sell_amt * exch_rate, 2) if sell_amt else None
        gp_usd = round(sell_amt - buy_amt, 2) if sell_amt and buy_amt else None
        gp_pct = round(gp_usd / sell_amt * 100, 2) if gp_usd and sell_amt and sell_amt != 0 else None
        gp_krw = round(gp_usd * exch_rate, 2) if gp_usd else None
        gp_pct_krw = round(gp_krw / sell_amt_krw * 100, 2) if gp_krw and sell_amt_krw and sell_amt_krw != 0 else None

        # 출고일자
        ship_date = row.get("출고일자")
        if isinstance(ship_date, pd.Timestamp):
            ship_date = ship_date.strftime("%Y-%m-%d")
        else:
            ship_date = str(ship_date) if ship_date and str(ship_date) != "nan" else None

        # Month
        month_val = row.get("출고년월")
        if month_val and str(month_val) != "nan":
            month_val = str(month_val).replace("/", "")
        else:
            month_val = None

        vendor = row.get("품목군") or row.get("품목대분류") or ""
        vendor = str(vendor) if str(vendor) != "nan" else ""

        records.append({
            "구분": vendor,
            "MPN": str(row.get("품번", "")) if str(row.get("품번", "")) != "nan" else "",
            "QTY": qty,
            "DCPL($)": dcpl,
            "매입금액($)": buy_amt,
            "SP($)": sp,
            "매출금액($)": sell_amt,
            "매출환율": exch_rate,
            "SP(KRW)": sp_krw,
            "매출금액(KRW)": sell_amt_krw,
            "GP($)": gp_usd,
            "GP%($)": gp_pct,
            "GP(KRW)": gp_krw,
            "GP%(KRW)": gp_pct_krw,
            "담당자": str(row.get("담당자", "")) if str(row.get("담당자", "")) != "nan" else "",
            "납품처": str(row.get("고객", "")) if str(row.get("고객", "")) != "nan" else "",
            "거래처코드": str(row.get("고객코드", "")) if str(row.get("고객코드", "")) != "nan" else "",
            "출고일자": ship_date,
            "입고일": inbound_date,
            "Month": month_val,
        })

    # DB 저장
    db.query(SalesModel).delete()
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]
    for r in records:
        db.add(sales_record_to_db(r, batch_id))
    db.commit()

    return {
        "columns": SALES_COLUMNS,
        "data": records,
        "total_rows": len(records),
        "summary": compute_sales_summary(records),
        "saved_to_db": True,
    }


@app.get("/api/sales/data")
async def get_sales_data(db: Session = Depends(get_db)):
    """DB에서 영업실적 데이터 조회"""
    rows = db.query(SalesModel).order_by(SalesModel.id).all()
    if not rows:
        return {"data": [], "columns": SALES_COLUMNS, "total_rows": 0}

    records = [sales_db_to_record(r) for r in rows]
    return {
        "columns": SALES_COLUMNS,
        "data": records,
        "total_rows": len(records),
        "summary": compute_sales_summary(records),
    }


@app.post("/api/reset-tables")
async def reset_tables():
    """DB 테이블 재생성 (스키마 변경 시)"""
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    return {"status": "ok", "message": "All tables recreated"}


# ==================== 마이크론 재고 ====================

MICRON_COLUMNS = [
    "Status", "Type", "PO", "DID", "MPN", "CPN (MOBIS ID 포함)", "BOX_TYPE",
    "QTY", "DNNo.", "Ship Date", "MicronInvoice#", "수입면장번호", "BL번호",
    "수입신고일", "FSE", "End customer", "Date Code",
    "Booking Customer & FSE", "Qty_booking", "비고",
]

micron_data = []  # 로컬 메모리 저장


@app.post("/api/micron/upload")
async def upload_micron(file: UploadFile = File(...)):
    global micron_data
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents), sheet_name="Detail", header=0)
    except Exception:
        return {"error": "Detail 시트를 찾을 수 없습니다."}

    records = []
    for idx, row in df.iterrows():
        rec = {}
        for col in df.columns:
            v = row[col]
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                rec[col] = None
            elif isinstance(v, pd.Timestamp):
                rec[col] = v.strftime("%Y-%m-%d")
            elif v is not None:
                rec[col] = str(v) if not isinstance(v, (int, float)) else v
            else:
                rec[col] = None
        rec["_id"] = str(idx)
        records.append(rec)

    micron_data = records

    return {
        "columns": MICRON_COLUMNS,
        "data": records,
        "total_rows": len(records),
    }


@app.get("/api/micron/data")
async def get_micron_data(
    status: str = "",
    did: str = "",
    mpn: str = "",
    notes_only: str = "",
):
    filtered = micron_data
    if status:
        filtered = [r for r in filtered if status in str(r.get("Status", ""))]
    if did:
        filtered = [r for r in filtered if did.lower() in str(r.get("DID", "")).lower()]
    if mpn:
        filtered = [r for r in filtered if mpn.lower() in str(r.get("MPN", "")).lower()]
    if notes_only == "true":
        filtered = [r for r in filtered if r.get("비고") and str(r.get("비고")) not in ("None", "", "nan")]

    return {"columns": MICRON_COLUMNS, "data": filtered, "total_rows": len(filtered)}


@app.get("/api/micron/summary/{did}")
async def micron_summary(did: str):
    """DID 기준 재고+입고예정 통합 조회"""
    items = [r for r in micron_data if str(r.get("DID", "")).upper() == did.upper()]
    if not items:
        return {"error": "해당 DID 없음"}

    by_status = {}
    for r in items:
        s = str(r.get("Status", "기타"))
        if s not in by_status:
            by_status[s] = {"count": 0, "qty": 0, "items": []}
        by_status[s]["count"] += 1
        by_status[s]["qty"] += to_float(r.get("QTY")) or 0
        by_status[s]["items"].append(r)

    total_qty = sum(v["qty"] for v in by_status.values())
    mpns = list(set(str(r.get("MPN", "")) for r in items))

    return {
        "did": did, "mpns": mpns, "total_qty": total_qty,
        "by_status": {k: {"count": v["count"], "qty": v["qty"]} for k, v in by_status.items()},
        "items": items,
    }


@app.post("/api/micron/update")
async def update_micron(data: dict):
    """R~T열 수정 (CS팀 권한)"""
    item_id = data.get("_id")
    if item_id is None:
        return {"error": "ID 없음"}

    for r in micron_data:
        if str(r.get("_id")) == str(item_id):
            r["Booking Customer & FSE"] = data.get("Booking Customer & FSE", r.get("Booking Customer & FSE"))
            r["Qty_booking"] = data.get("Qty_booking", r.get("Qty_booking"))
            r["비고"] = data.get("비고", r.get("비고"))
            return {"updated": item_id}

    return {"error": "not found"}


# ==================== 거래명세서 ====================

import requests as http_requests

@app.get("/api/exchange-rate")
async def get_exchange_rate():
    """실시간 USD/KRW 환율 조회"""
    try:
        r = http_requests.get("https://open.er-api.com/v6/latest/USD", timeout=5)
        rate = r.json()["rates"]["KRW"]
        return {"rate": round(rate, 2), "source": "exchangerate-api"}
    except Exception:
        return {"rate": 1400, "source": "fallback"}


@app.post("/api/invoice/generate")
async def generate_invoice(data: dict):
    """거래명세서 엑셀 생성 — 원본 동일 형식 + 도장"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage

    items = data.get("items", [])
    customer = data.get("customer", "")
    date_str = data.get("date", "")
    rate = float(data.get("rate", 1400))

    wb = Workbook()
    ws = wb.active
    ws.title = "거래명세서"

    # 열 너비 (원본 동일)
    widths = {"A": 5, "B": 20, "C": 8, "D": 14, "E": 16, "F": 10, "G": 16, "H": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 스타일
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right_a = Alignment(horizontal="right", vertical="center")
    left_a = Alignment(horizontal="left", vertical="center")

    # Row 1: 거래명세표
    ws.merge_cells("A1:H1")
    ws["A1"] = "거래명세표"
    ws["A1"].font = Font(bold=True, size=36)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 55
    ws.row_dimensions[2].height = 8

    # 굵은 테두리
    thick = Side(style="medium")
    thick_border_tl = Border(left=thick, top=thick)
    thick_border_t = Border(top=thick)
    thick_border_tr = Border(right=thick, top=thick)
    thick_border_l = Border(left=thick)
    thick_border_r = Border(right=thick)
    thick_border_bl = Border(left=thick, bottom=thick)
    thick_border_b = Border(bottom=thick)
    thick_border_br = Border(right=thick, bottom=thick)
    no_border = Border()

    # Row 3: 공급자 제목
    ws.merge_cells("A3:C3")
    ws["A3"] = "공 급 자"
    ws["A3"].font = Font(bold=True, size=10)
    ws["A3"].alignment = center
    ws["A3"].border = Border(left=thick, top=thick, bottom=Side(style="thin"))
    ws["B3"].border = Border(top=thick, bottom=Side(style="thin"))
    ws["C3"].border = Border(right=thick, top=thick, bottom=Side(style="thin"))

    # Row 4~9: 공급자 정보 (A~C, 외곽 굵은 테두리, 내부 없음)
    info = [
        (4, "등록번호 : 229-81-00105"),
        (5, "상      호 : ㈜유니트론텍"),
        (6, "대표이사 : 남궁 선"),
        (7, "주 : 서울 강남구 영동대로 638(삼성동, 삼보빌딩 9층)"),
        (8, "업      태 : 도.소매"),
        (9, "종      목 :전자부품 외"),
    ]
    for row_num, val in info:
        ws.merge_cells(f"A{row_num}:C{row_num}")
        ws[f"A{row_num}"] = val
        ws[f"A{row_num}"].font = Font(size=8)
        ws[f"A{row_num}"].alignment = left_a
        # 외곽만 굵게
        if row_num == 9:  # 마지막 행
            ws[f"A{row_num}"].border = Border(left=thick, bottom=thick)
            ws[f"B{row_num}"].border = Border(bottom=thick)
            ws[f"C{row_num}"].border = Border(right=thick, bottom=thick)
        else:
            ws[f"A{row_num}"].border = Border(left=thick)
            ws[f"C{row_num}"].border = Border(right=thick)

    # 공급받는자 제목
    ws.merge_cells("F3:H3")
    ws["F3"] = "공급받는자"
    ws["F3"].font = Font(bold=True, size=10)
    ws["F3"].alignment = center
    ws["F3"].border = Border(left=thick, top=thick, bottom=Side(style="thin"))
    ws["G3"].border = Border(top=thick, bottom=Side(style="thin"))
    ws["H3"].border = Border(right=thick, top=thick, bottom=Side(style="thin"))

    # 공급받는자 본문 (F4:H9 전체 병합, 굵은 외곽)
    ws.merge_cells("F4:H9")
    ws["F4"] = customer
    ws["F4"].font = Font(bold=True, size=16)
    ws["F4"].alignment = center
    # 병합 셀 외곽 굵은 테두리
    for row_num in range(4, 10):
        for col_letter in ["F", "G", "H"]:
            cell = ws[f"{col_letter}{row_num}"]
            l = thick if col_letter == "F" else None
            r_side = thick if col_letter == "H" else None
            b = thick if row_num == 9 else None
            cell.border = Border(
                left=l or Side(), right=r_side or Side(),
                bottom=b or Side(), top=Side(),
            )

    # Row 12: 날짜
    ws["B12"] = date_str
    ws["B12"].font = Font(bold=True, size=10)

    # Row 13: 헤더
    headers = ["No.", "Part #", "QTY", "U/PRICE ($)", "Amount ($)", "RATE", "U/PRICE (￦)", "AMOUNT (￦)"]
    header_fill = PatternFill(start_color="CCFF33", end_color="CCFF33", fill_type="solid")
    for j, h in enumerate(headers):
        cell = ws.cell(row=13, column=j+1)
        cell.value = h
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # 데이터 행
    total_usd = 0
    total_krw = 0
    for i, item in enumerate(items):
        row = 14 + i
        qty = float(item.get("qty", 0))
        price_usd = float(item.get("price", 0))
        amount_usd = round(qty * price_usd, 2)
        price_krw = round(price_usd * rate, 2)
        amount_krw = round(amount_usd * rate, 0)
        total_usd += amount_usd
        total_krw += amount_krw

        ws.cell(row=row, column=1, value=i+1).alignment = center
        ws.cell(row=row, column=2, value=item.get("part", "")).alignment = center
        ws.cell(row=row, column=3, value=int(qty)).alignment = center
        c4 = ws.cell(row=row, column=4)
        c4.value = price_usd
        c4.number_format = '$#,##0.00'
        c4.alignment = center
        c5 = ws.cell(row=row, column=5)
        c5.value = amount_usd
        c5.number_format = '$#,##0.00'
        c5.alignment = center
        c6 = ws.cell(row=row, column=6)
        c6.value = rate
        c6.number_format = '#,##0.00'
        c6.alignment = center
        c7 = ws.cell(row=row, column=7)
        c7.value = price_krw
        c7.number_format = '₩#,##0.00'
        c7.alignment = center
        c8 = ws.cell(row=row, column=8)
        c8.value = int(amount_krw)
        c8.number_format = '₩#,##0'
        c8.alignment = center

        for j in range(1, 9):
            ws.cell(row=row, column=j).border = border

    # 빈 행 (10행까지)
    for i in range(len(items), 10):
        row = 14 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center
        for j in range(1, 9):
            ws.cell(row=row, column=j).border = border

    # 소계/부가세/합계
    summary_row = 24
    tax_usd = round(total_usd * 0.1, 2)
    tax_krw = round(total_krw * 0.1, 0)
    for label, usd_val, krw_val, row_offset in [
        ("소  계", total_usd, total_krw, 0),
        ("부가세", tax_usd, tax_krw, 1),
        ("합  계", total_usd + tax_usd, total_krw + tax_krw, 2),
    ]:
        r = summary_row + row_offset
        fnt = Font(bold=True, size=9)

        ws.cell(row=r, column=4, value=label).font = fnt
        ws.cell(row=r, column=4).alignment = center
        c5 = ws.cell(row=r, column=5)
        c5.value = usd_val
        c5.number_format = '$#,##0.00'
        c5.alignment = center
        c5.font = fnt
        ws.cell(row=r, column=7, value=label).font = fnt
        ws.cell(row=r, column=7).alignment = center
        c8 = ws.cell(row=r, column=8)
        c8.value = int(krw_val)
        c8.number_format = '₩#,##0'
        c8.alignment = center
        c8.font = fnt

        for j in [4, 5, 7, 8]:
            ws.cell(row=r, column=j).border = border

    # 합계 아래 줄 (A~H 전체)
    합계row = summary_row + 2
    for j in range(1, 9):
        cell = ws.cell(row=합계row, column=j)
        existing = cell.border
        cell.border = Border(
            left=existing.left, right=existing.right,
            top=existing.top, bottom=Side(style="medium"),
        )

    # 비고
    ws.merge_cells("A27:H27")
    ws["A27"] = "  비  고 : 금일 최초고시 매매기준율 적용"
    ws["A27"].font = Font(size=9)

    # 인수자 + 아래 줄
    ws.merge_cells("E30:H30")
    ws["E30"] = "인수자 :                                              (인)"
    for j in range(1, 9):
        ws.cell(row=30, column=j).border = Border(bottom=Side(style="thin"))
    ws["E30"].font = Font(size=9)

    # 계좌정보
    ws["B31"] = "계좌정보"
    ws["B31"].font = Font(bold=True, size=9)
    ws["B32"] = "원화> 기업은행 528-002245-01011"
    ws["B32"].font = Font(size=9)
    ws["B33"] = "외화> 기업은행 528-002245-56-00013"
    ws["B33"].font = Font(size=9)

    # 도장 이미지
    stamp_path = os.path.join(os.path.dirname(__file__), "stamp.png")
    if os.path.exists(stamp_path):
        img = XlImage(stamp_path)
        from openpyxl.utils.units import pixels_to_EMU
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
        img.width = 75
        img.height = 75
        # 공급자 박스 우측 상단에 겹치게 — C열, 2행에서 시작 (오프셋으로 미세 조정)
        marker = AnchorMarker(col=2, colOff=300000, row=1, rowOff=50000)  # C열 중간, 2행 상단
        marker2 = AnchorMarker(col=3, colOff=200000, row=4, rowOff=100000)
        anchor = TwoCellAnchor(_from=marker, to=marker2)
        img.anchor = anchor
        ws.add_image(img)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=invoice_{date_str}.xlsx"},
    )


# 프론트엔드 정적 파일 서빙 (배포용)
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=os.path.join(STATIC_DIR, "static")), name="static-assets")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        file_path = os.path.join(STATIC_DIR, full_path)
        if full_path and os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(STATIC_DIR, "index.html"))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
